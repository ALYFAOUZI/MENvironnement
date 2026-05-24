import pandas as pd

#worksheet = pd.read_excel('Sys_Fonc_Travail.xlsx', sheet_name = 'Fonc_Aff')
#print(worksheet['B4'].iloc[0])
#print(worksheet)
import re # for regular expression
import openpyxl 
import xlrd
'''
wb = openpyxl.load_workbook('Sys_Fonc_Travail.xlsx')
ws = wb.active
#ws = wb['vgsales'] can use worksheet name
print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
#print('The value in cell A1 is: '+ws['A2'].value)
data=[ws.cell(row=i,column=1).value for i in range(2,12)]
print(data)
# reading data from a range of cells (from column 1 to 6)

my_list = list()

for value in ws.iter_rows(
    min_row=1, max_row=11, min_col=1, max_col=6, 
    values_only=True):
    my_list.append(value)
'''

excel_data_df = pd.read_excel('Sys_Fonc_Travail.xlsx', sheet_name='Fonc_Aff',index_col =None,header=None)
# print whole sheet data
#print(excel_data_df.columns)
#print(excel_data_df.iloc[0,4]  )
#for i in [excel_data_df[j][k] for k in range(0,len(excel_data_df)) for j in excel_data_df.columns]:
       #print(i)
print(len(excel_data_df))
data_result = []
for k in range(0,len(excel_data_df)):
       for j in excel_data_df.columns:
           
         if (excel_data_df[j][k] == 'X' or not excel_data_df.isnull()[j][k] ) and excel_data_df[j][k] != 'nan'  and not excel_data_df.isnull()[2][k] :
         #'AP - Aires protégées':
            # and excel_data_df[j][0]== 'AP - Aires protégées'
            #and excel_data_df[j][0]== 'CLIMATO - Climatologie (collecte et diffusion)'
            #print(j,k)
            #print(excel_data_df[j][k])
            #print(excel_data_df.astype(str)[0][k] + '_' + excel_data_df.astype(str)[1][k] + '_'+ excel_data_df[2][k] )
            #print(excel_data_df[j][0])
            #str = excel_data_df.astype(str)[0][k] + '_' + excel_data_df.astype(str)[1][k] + '_'+ excel_data_df[2][k]
            #if any(c.isalpha() for c in str):
            #if re.search('[a-zA-Z]+',str):
            if not excel_data_df.isnull()[j][0] and excel_data_df.astype(str)[j][0]  !=' ' and excel_data_df.astype(str)[0][k] + '_' + excel_data_df.astype(str)[1][k] != 'nan_nan':


              data_result.append(
              {'Domain':  excel_data_df.astype(str)[0][k] + '_' + excel_data_df.astype(str)[1][k] + '_'+ excel_data_df[2][k] ,
                     'System':excel_data_df[j][0] }
              )
df = pd.DataFrame(data_result)
S_FILE_OUTPUT = 'C:\\Users\\foual05\\Prototype_STH\\Test.csv' 
df.to_csv(S_FILE_OUTPUT, sep='|', header=True, index=False)